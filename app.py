import json
import openpyxl
from datetime import datetime, timedelta
from pandas import DataFrame
import random
import xlsxwriter

def XlsxMake(final_schedule, team1, team2):
    title = team1 + " vs " + team2
    # for i in final_schedule.keys():
    #    title = title + " vs " + i
    court_counter = len(final_schedule[list(final_schedule.keys())[0]].keys())
    print(court_counter)

    workbook = xlsxwriter.Workbook('result.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_column("A:M", 15)
    worksheet.set_column("E:F", 20)
    worksheet.set_column("H:I", 20)
    meet_title_format = workbook.add_format(
        {
            "bold": 1,
            "underline" : 1,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "fg_color": "#FFD24C",
            "font" : "montserrat",
            "size" : 22,
            
        }
    )
    categories = workbook.add_format( 
        {   
            "bold": 1,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "color" : "#FFFFFF",
            "fg_color": "#355B85",
            "font" : "montserrat",
            "size" : 10,
            }
    )
    data_blue = workbook.add_format(
        {
        "font" : "montserrat",
        "align": "center",
        "size" : 10,
        "border": 1,
        "fg_color": "#B3C1D1",
        }
    )
    data_yellow = workbook.add_format(
        {
        "font" : "montserrat",
        "align": "center",
        "size" : 10,
        "border": 1,
        "fg_color": "#FFDF80",
        }
    )
    data_black = workbook.add_format(
        {
        "font" : "montserrat",
        "align": "center",
        "size" : 10,
        "border": 1,
        "bold" : 1,
        "color" : "#FFFFFF",
        "fg_color": "#333333",
        }
    )
    data_red = workbook.add_format(
        {
        "font" : "montserrat",
        "align": "center",
        "size" : 10,
        "border": 1,
        "bold" : 1,
        "color" : "#FFFFFF",
        "fg_color": "#79242F",
        }
    )
    data_grey = workbook.add_format(
        {
        "font" : "montserrat",
        "align": "center",
        "size" : 10,
        "border": 1,
        "fg_color": "#CCCCCC",
        }
    )
    print(final_schedule)
    worksheet.merge_range("A1:K4", title, meet_title_format)

    worksheet.write('A5', 'Schedule', categories)
    worksheet.write('B5', 'Court', categories)
    worksheet.write('C5', 'In progress', categories)
    worksheet.write('D5', 'Event', categories)
    worksheet.merge_range('E5:F5', team1, categories)
    worksheet.write('G5', 'vs.', categories)
    worksheet.merge_range('H5:I5', team2, categories)
    worksheet.merge_range('J5:K5', 'Score (Winner First)', categories)
    start_counter = 6
    start_cell = "A6"
    end_cell = "A6"
    team1_ATeam = []
    team2_ATeam = []
    team1_OTeam = []
    team2_OTeam = []
    for sched_i in final_schedule:
        end_counter = start_counter + court_counter -1
        start_cell = "A" + str(start_counter)
        end_cell=  "A" + str(end_counter)
        t_format = start_cell + ":" + end_cell
        if final_schedule[sched_i] != {}:
            worksheet.merge_range(t_format, sched_i, categories)
        c_number = 1
        for i in range(int(start_counter), int(end_counter)+1):
            court_cell = "B"+ str(i)
            event_cell = "D" + str(i)
            vs_cell = "G" + str(i)
            progress_cell = "C" + str(i)
            team1_cell = "E"+ str(i) + ":F" + str(i)
            team2_cell = "H" + str(i) + ":I" + str(i)
            score_cell = "J" + str(i) + ":K" + str(i)
            victory_cell_t1 = "L" + str(i)
            victory_cell_t2 = "M" + str(i)
            try:
                #print(final_schedule[sched_i][str(c_number)][1])
                event_data = final_schedule[sched_i][c_number][0]
                if (event_data[2] == "1" or event_data[2] == "2" or event_data[2] == "3") and len(event_data) == 3:
                    #print("this is a priority match", event_data)
                    team1_ATeam.append("L" + str(i))
                    team2_ATeam.append("M" + str(i))
                worksheet.write(progress_cell, "", data_yellow)
                try:
                    team1_data = final_schedule[sched_i][c_number][1][0] + " + " + final_schedule[sched_i][c_number][1][1]
                except:
                    team1_data = final_schedule[sched_i][c_number][1][0]
                try:
                    team2_data = final_schedule[sched_i][c_number][2][0] + " + " + final_schedule[sched_i][c_number][2][1]
                except:
                    team2_data = final_schedule[sched_i][c_number][2][0]
                worksheet.write(event_cell, event_data, data_yellow)
                worksheet.write(vs_cell, "vs", categories)
                worksheet.merge_range(team1_cell, str(team1_data), data_blue)
                worksheet.merge_range(team2_cell, str(team2_data), data_blue)
                worksheet.merge_range(score_cell, "",data_yellow)
                worksheet.write(victory_cell_t1, 1, data_blue)
                worksheet.write(victory_cell_t2, 1, data_blue)
                worksheet.write(court_cell, c_number, data_blue)
                team1_OTeam.append("L" + str(i))
                team2_OTeam.append("M" + str(i))

            
            except:
                pass
            c_number +=1
        start_counter = end_counter+1
    worksheet.write("L1", team1, data_red)
    worksheet.write("M1", team2, data_red)
    worksheet.merge_range("L2:M2", "A-Team Tally", data_black)
    worksheet.merge_range("L4:M4", "Overall Tally", data_black)
    sumat1 = ""
    sumat2 = ""
    sumot1 = ""
    sumot2 = ""
    for at in range(0, len(team1_ATeam)):
        sumat1 = sumat1 + ", " + team1_ATeam[at]
        sumat2 = sumat2 + ", " + team2_ATeam[at]
    for ot in range(0, len(team1_OTeam)):
        sumot1 = sumot1 + ", " + team1_OTeam[ot]
        sumot2 = sumot2 + ", " + team2_OTeam[ot]
    sumat1 = "=SUM(" + sumat1[1:] + ")"
    sumat2 = "=SUM(" + sumat2[1:] + ")"
    sumot1 = "=SUM(" + sumot1[1:] + ")"
    sumot2 = "=SUM(" + sumot2[1:] + ")"
    worksheet.write_formula('L3', sumat1, data_grey)
    worksheet.write_formula('M3', sumat2, data_grey)
    worksheet.write_formula('L5', sumot1, data_grey)
    worksheet.write_formula('M5', sumot2, data_grey)
    print(sumat1)
        #worksheet.write(event_cell, final_schedule[sched_i][str(c_number)][0])
            
    workbook.close()
def selectMatch(match_number, Meet, players, timeslot, prio_flag, priority_courts, the_court_number, priority_dict, conflicts_dict):
   prev_player = ""
   flag = False
   conflict = False

   while True: #to make sure same matches are no longer called
      if match_number == {}:
         return True
      
      if prio_flag == True and the_court_number <= int(priority_courts):
         prio_list = []
         for i in Meet:
            if Meet[i] == False:
               prio_list.append(i)
         if len(prio_list) == 0:
            match = random.choice(list(match_number.keys()))
         else:
            match = random.choice(prio_list)
      else:
         match = random.choice(list(match_number.keys()))

      if prio_flag == True:
         if match_number[match][1] == True and match_number[match][2] == True and match_number[match][3] == True:
            priority_dict[match] = True
      
      if False not in match_number[match].values():
         del match_number[match]
         continue

      if prio_flag == True and the_court_number <= int(priority_courts):
         if priority_dict[match] == False:
            rank_choice = random.choice([1, 2, 3])
         elif len(prio_list) != 0:
            temp_list = []
            for i in range(4, len(match_number[match].keys())+1):
               temp_list.append(i)
            rank_choice = random.choice(temp_list)
         else:
            rank_choice = random.choice(list(match_number[match].keys()))
      else:
         rank_choice = random.choice(list(match_number[match].keys()))

      if match_number[match][rank_choice] == True:
         continue
      else:
         #conflict check
         for i in Meet:
            rank_number = "Rank " + str(rank_choice)

            try:
               for j in range(0, len(Meet[i][match][rank_number]["Player Name"])):
                  temp_player = Meet[i][match][rank_number]["Player Name"][j]
                  if temp_player in players:
                     conflict = True
                     break
            except:
               temp_num = rank_choice

               while temp_num > len(Meet[i][match].values()):   
                  temp_num = abs(len(Meet[i][match].values()) - temp_num)

               rank_number = "Rank " + str(temp_num)
               for j in range(0, len(Meet[i][match][rank_number]["Player Name"])):
                  temp_player = Meet[i][match][rank_number]["Player Name"][j]

                  if temp_player in players:
                     conflict = True
                     break

         if conflict == True and prev_player != temp_player:
            prev_player = temp_player
            conflict = False
            continue
         else:
            if prev_player == temp_player:
               print("there will be an unavoidable conflict in time", timeslot, prev_player)
               if timeslot not in conflicts_dict:
                  conflicts_dict[timeslot] = [prev_player]
               else:
                  conflicts_dict[timeslot].append(prev_player)
            match_number[match][rank_choice] = True
            break
   return [match,rank_choice]
         
def menu():
    print("----------------------Menu----------------------\n" + 
          " [Add Roster -> [A]]"+
          "\n [View Roster -> [V]]"+
          "\n [Save Changes -> [S]]"+
          "\n [Make Meet -> [M]]"+
          "\n [Export to XLSX -> [E]]"+ 
          "\n [Terminate -> [X]]")
    
    return input("\n<Your Choice>")

#may not need:   
def save(Meet, text_file):
    json_format = json.dumps(Meet, indent=3)

    file_to_delete = open(text_file,'w')
    file_to_delete.close()

    with open(text_file, 'w') as convert_file: 
        convert_file.write(json_format)
    print("saved final schedule!")

def makeMeet(Meet):# for a MEET, not trimeet
   #first make structure: how long each game will take, how many courts, how much time
   #priority courts named courts 1, 2 and 3. <- these hold the best courts.
   print("//MAKE PROGRAM START\\\ ")
   prio_flag = False
   #-- a series of inputs to begin the program

   tpm = input("what is the time given for each game cycle? write in terms of minutes(30 = 30 minutes)") #tpm = time per match
   meet_time_start = input("when does the meet start? Write in the 24:00 clock format(16:00 = 4pm)")
   meet_time_end = input("When does the meet end? Write in the 24:00 clock format(16:00 = 4pm)")
   courts_number = input("How many courts are used in this meet?")
   priority_query = input("Do you want to have priority courts?\nPriority courts are for the top seeds[1 - 3] to play on certain courts.\n[Y/N]")
   if priority_query.upper() == "Y":
      priority_courts = input("How many priority courts would you like?")
      if priority_courts >= courts_number:

         print("You cannot have equal or more priority courts than normal courts")
         return
      print("//PRIORITY COURTS = COURTS 1 -", priority_courts)
      prio_flag = True
   else:
      print("Running program without priority courts.")
   #-- using the datetime package to be ale to transfer str to time, and do math with time.

   time_begin = datetime.strptime(meet_time_start, "%H:%M")#check times
   time_end = datetime.strptime(meet_time_end, "%H:%M")
   difference = time_end - time_begin
   plausible_minutes = difference.total_seconds() / 60 #calculates minutes alloted to the meet
   plausible_total_matches = (plausible_minutes / int(tpm)) * int(courts_number) #minutes over time per match to get amount of matches for 1 court, * court no.
   match_number = { #will be changed later
      "MD" : 0,
      "MS" : 0,
      "XD" : 0,
      "WD" : 0,
      "WS" : 0
   }
   for i in Meet: #searches and finds the number of matches that will occur, from the highest number of spots from each event of each time.
      for j in Meet[i]:
         if len(Meet[i][j].values()) > match_number[j]:
               match_number[j] = len(Meet[i][j].values()) #places the values into match_number for calculation later.
   total_matches = sum(match_number.values()) # number of matches that have to be played

   #-- a check with time. If plausible with the time given and matches that have to be played, will allow, if not, display why not and break.
   
   print("PLAUSible matches", plausible_total_matches, "||TOTAL matches" , total_matches)
   if(plausible_total_matches < total_matches): #if amount of matches > matches can be played, impossible to create schedule
      print("NO TIME FOR THE MEET\nPlausible matches based on tpm and meet time:",plausible_total_matches, "\nTotal matches required for meet:", total_matches)
      return
   
   #--begin final part
   final_schedule = dict()#final schedule to be given out.
   priority_dict = { "MD" : False,
                     "MS" : False,
                     "XD" : False,
                     "WD" : False,
                     "WS" : False
                   }
   conflicts = dict()
   time_change = timedelta(minutes = int(tpm))
   new_time = time_begin
   while new_time < time_end:
       final_schedule[new_time.strftime('%I:%M')] = dict()
       new_time = new_time + time_change

   for c in match_number: #repurpose match_number
      temp = int(match_number[c])
      match_number[c] = dict()
      for d in range(1, temp + 1):
            match_number[c][d] = False

   for timeslot in final_schedule.keys():
      courts = dict()
      players = []
      for i in range(1, int(courts_number)+1):
         if match_number == {}:
            break
         SM_Results = selectMatch(match_number, Meet, players, timeslot, prio_flag, priority_courts, i, priority_dict, conflicts)
         if SM_Results == True:
            break
         match = SM_Results[0]
         rank_choice = SM_Results[1]
                                                                                                                                                
         courts[i] = []
         event_temp = match + str(rank_choice)
         courts[i].append(event_temp)

         for a in Meet:
            rank_number = "Rank " + str(rank_choice)
            try:
               courts[i].append(Meet[a][match][rank_number]["Player Name"])
            except:
               temp_num = rank_choice
               while temp_num > len(Meet[a][match].values()):   
                  temp_num = abs(len(Meet[a][match].values()) - temp_num)
               rank_number = "Rank " + str(temp_num)
               courts[i].append(Meet[a][match][rank_number]["Player Name"])

            x = 0
            while x < len(Meet[a][match][rank_number]["Player Name"]):
               players.append(Meet[a][match][rank_number]["Player Name"][x])
               x+=1
      final_schedule[timeslot] = courts
   save(final_schedule, "schedule.txt")
   return final_schedule
# -> addEvent(Meet) :: add md, ms, xd, ws, and wd events to the algorithm.   
def addEvent(Meet): 
    for i in Meet:
        Meet[i]["MD"] = dict()
        Meet[i]["MS"] = dict()
        Meet[i]["XD"] = dict()
        Meet[i]["WS"] = dict()
        Meet[i]["WD"] = dict()
# -> addPlayer() :: by inputting the player's name, event number, event title, and ranks of those events, input them into the json file.
def addPlayer(Meet, team_name, players):
    file_name = input("What is the file name called?")
    excel_sheet = openpyxl.load_workbook(file_name)

    sh = excel_sheet.active
    temp = []
    for i in range(1, sh.max_row+1): 
        for j in range(1, 4): 
            cell_obj = sh.cell(row=i, column=j)
            if cell_obj.value == None:
                pass
            else:
                temp.append(cell_obj.value)
    i = 0
    while True:
        if i >= len(temp):
            break
        if temp[i] == "MD" or temp[i] == "MS" or temp[i] == "XD" or temp[i] == "WS" or temp[i] == "WD":
            curr_event = temp[i]
        elif temp[i] != str(temp[i]):
            json_rank_text = "Rank " + str(int(temp[i]))
        else:
            name = temp[i]
            if json_rank_text not in Meet[curr_event].keys(): # if empty
                 Meet[curr_event][json_rank_text] = {"Player Name" : [name]}
            else:
                 Meet[curr_event][json_rank_text]["Player Name"].append(name)
        i+=1



teams = [] # a list of teams.
teamNumber = input("//PROGRAM START\\\ \nHow many teams are there in this meet?")
players = dict() # a dictionary of players relevant to each team.
#---
Meet = dict()
for i in range(1, int(teamNumber) + 1):
    print("--------------------------")
    temp = input("What is the team name of team " +  str(i) + "?")
    temp = temp.upper()
    Meet[temp] = dict()
    teams.append(temp)
#---- add the events within the team
addEvent(Meet)
#----
if int(teamNumber) == 2:
    print("--------------------------")
    print("Welcome to the meet between", list(Meet.keys())[0],"and", list(Meet.keys())[1] + "!")
else:
    print("--------------------------")
    print("Welcome to the Trimeet between", list(Meet.keys())[0]+",", list(Meet.keys())[1],"and", list(Meet.keys())[2] + "!")
#----gonna have to loop the process below----
fin_sched = dict()
while True:
    menu_query = menu()
    #try:
    if menu_query.upper() == "A":
        temp = input("Selection : Add Player : Which team is this player on?")
        if temp.upper() == list(Meet.keys())[0]:
            addPlayer(Meet[teams[0]], teams[0], players)
        elif temp.upper() == list(Meet.keys())[1]:
            addPlayer(Meet[teams[1]], teams[1], players)
        elif len(teams) > 2:
            if temp.upper() == list(Meet.keys())[2]:
                addPlayer(Meet[teams[2]], teams[2], players)
        else:
            print("Not a team!")
    elif menu_query.upper() == "S":
        save(Meet, "save.txt")
    elif menu_query.upper() == "V":
        print("this function has not been implemented yet")
        break
    elif menu_query.upper() == "M":
        fin_sched = makeMeet(Meet)
    elif menu_query.upper() == "E" and fin_sched != {}:
        XlsxMake(fin_sched, list(Meet.keys())[0], list(Meet.keys())[1])
    elif menu_query.upper() == "X":
            break
    else:
        print("something wrong occured")
    #except:
        #print("not the right answer!")# incase something is inputted wrong