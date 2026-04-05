import json
import openpyxl
from datetime import datetime, timedelta
from pandas import DataFrame
import random
import sys
import tty
import termios


BOLD = '\033[1m'
END = '\033[0m'

def get_key():
    """Get a single keypress from the user."""
    fd = sys.stdin.fileno()
    old_settings = termios.tcgetattr(fd)
    try:
        tty.setraw(sys.stdin.fileno())
        ch = sys.stdin.read(1)
        # Handle arrow keys (they send 3 characters)
        if ch == '\x1b':
            next_chars = sys.stdin.read(2)
            if next_chars == '[A':
                return 'UP'
            elif next_chars == '[B':
                return 'DOWN'
            return None
        elif ch == '\r':  # Carriage return
            return 'ENTER'
        elif ch == '\n':  # Newline
            return 'ENTER'
        elif ch == ' ':  # Spacebar as alternative
            return 'ENTER'
        elif ch == '\x03':  # Ctrl+C
            return 'QUIT'
        elif ch == 'q' or ch == 'Q':  # Q to quit
            return 'QUIT'
        return None
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)

# -> addEvent(Meet), add md, ms, xd, ws, and wd events to the algorithm.
def addEvent(Meet): 
    for i in Meet:
        Meet[i]["MD"] = dict()
        Meet[i]["MS"] = dict()
        Meet[i]["XD"] = dict()
        Meet[i]["WS"] = dict()
        Meet[i]["WD"] = dict()

# -> addPlayer(), by inputting the player's name, event number, event title, and ranks of those events, input them into the json file.
def addPlayer(Meet, team_name, players):
    file_name = input("What is the file name called?")
    excel_sheet = openpyxl.load_workbook(file_name)

    sh = excel_sheet.active
    temp = []
    for i in range(1, sh.max_row+1): 
        for j in range(1, 4): 
            cell_obj = sh.cell(row=i, column=j)
            if cell_obj.value == None:
                pass
            else:
                temp.append(cell_obj.value)
    i = 0
    while True:
        if i >= len(temp):
            break
        if temp[i] == "MD" or temp[i] == "MS" or temp[i] == "XD" or temp[i] == "WS" or temp[i] == "WD":
            curr_event = temp[i]
        elif temp[i] != str(temp[i]):
            json_rank_text = "Rank " + str(int(temp[i]))
        else:
            name = temp[i]
            if json_rank_text not in Meet[curr_event].keys(): # if empty
                 Meet[curr_event][json_rank_text] = {"Player Name" : [name]}
            else:
                 Meet[curr_event][json_rank_text]["Player Name"].append(name)
        i+=1

def menu():
    options = [
        ("Add Roster", "A"),
        ("Save", "S"),
        ("Make Meet", "M"),
        ("Terminate", "X")
    ]
    
    selected = 0
    num_options = len(options)
    
    # Print static header once
    print("\n" + "="*50)
    print(f"{BOLD}{'BADMINTON MEET MAKER - MAIN MENU':^50}{END}")
    print("="*50)
    print(f"\n  Use ↑/↓ arrows to navigate, Enter to select\n")
    print("  " + "-"*46)
    
    def draw_menu():
        """Redraw the menu options."""
        for i, (option, command) in enumerate(options):
            if i == selected:
                print(f"  {BOLD}→ {option:<28}{END} {BOLD}[{command}]{END}")
            else:
                print(f"    {option:<28} [{command}]")
        print(f"  {'-'*46}")
        sys.stdout.flush()
    
    # Draw initial menu
    draw_menu()
    
    while True:
        key = get_key()
        
        if key == 'UP':
            selected = (selected - 1) % num_options
            # Move cursor up to redraw menu (num_options + 1 for separator line)
            print(f"\033[{num_options + 1}A", end='')
            draw_menu()
            
        elif key == 'DOWN':
            selected = (selected + 1) % num_options
            # Move cursor up to redraw menu
            print(f"\033[{num_options + 1}A", end='')
            draw_menu()
            
        elif key == 'ENTER':
            print("\n")  # Add some space before returning
            return options[selected][1]  # Return the command number
            
        elif key == 'QUIT':
            print("\n")
            return '3'  # Return terminate option
        

def saveSchedule(Meet):
    json_format = json.dumps(Meet, indent=3)

    file_to_delete = open("schedule.txt",'w')
    file_to_delete.close()

    with open('schedule.txt', 'w') as convert_file: 
        convert_file.write(json_format)
    print("saved!")

def save(Meet):
    json_format = json.dumps(Meet, indent=3)

    file_to_delete = open("save.txt",'w')
    file_to_delete.close()

    with open('save.txt', 'w') as convert_file: 
        convert_file.write(json_format)
    print("saved!")
# ----------------------------------
# **********************************
# *********START OF PROGRAM*********
# **********************************
# ----------------------------------

teams = [] # Array of Teams
players = dict() # a dictionary of players relevant to each team.
Meet = dict()
#/----------
#Initial Print Statement
print("\n")
print("="*65)
print(f"{BOLD}{'BADMINTON MEET MAKER':^65}{END}")
print(f"{'by Richard Huang and BUCD':^65}")
print("="*65)
print("\n" + " "*5 + "This program is designed to help create a 1v1 school badminton")
print(" "*5 + "meet, generating a conflict-free .xlsx file for users.")
print(" "*5 + "Just answer the prompts and follow the instructions.")
print("\n" + "="*65 + "\n")
#\----------
for i in range(1,3):
    print("_______________________________________________________________")
    temp = input(f"{BOLD}What is the team name of team {i}?{END}>> ")
    temp = temp.upper()
    Meet[temp] = dict()
    teams.append(temp)
#---- add the events within the team
addEvent(Meet)
#----
print("--------------------------")
print("Welcome to the meet between", list(Meet.keys())[0],"and", list(Meet.keys())[1] + "!")

#----gonna have to loop the process below----
while True:
    menu_query = menu()
    #try:
    if menu_query.upper() == "A":
        temp = input("Selection == ")
        if temp.upper() == list(Meet.keys())[0]:
            addPlayer(Meet[teams[0]], teams[0], players)
        elif temp.upper() == list(Meet.keys())[1]:
            addPlayer(Meet[teams[1]], teams[1], players)
        elif len(teams) > 2:
            if temp.upper() == list(Meet.keys())[2]:
                addPlayer(Meet[teams[2]], teams[2], players)
        else:
            print("Not a team!")
    if menu_query.upper() == "S":
        save(Meet)
    if menu_query.upper() == "V":
        print("this function has not been implemented yet")
        break
    if menu_query.upper() == "M":
        pass
        #makeMeet(Meet)
    if menu_query.upper() == "X":
         break
    
print(Meet)