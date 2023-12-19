#import string
#PARA
import json
import openpyxl
import datetime
#should keep a list of all players inputted, so that we could begin doing conflicts too.
# perhaps instead of writing team every time, can set just 1 team everytime.
# -> addEvent(Meet) :: add md, ms, xd, ws, and wd events to the algorithm.
def addEvent(Meet): 
    for i in Meet:
        Meet[i]["MD"] = dict()
        Meet[i]["MS"] = dict()
        Meet[i]["XD"] = dict()
        Meet[i]["WS"] = dict()
        Meet[i]["WD"] = dict()
# -> addPlayer() :: by inputting the player's name, event number, event title, and ranks of those events, input them into the json file.
def addPlayer(Meet, team_name, players):
    file_name = input("What is the file name called?")
    excel_sheet = openpyxl.load_workbook(file_name)

    sh = excel_sheet.active
    temp = []
    for i in range(1, sh.max_row+1): 
        for j in range(1, 4): 
            cell_obj = sh.cell(row=i, column=j)
            if cell_obj.value == None:
                pass
            else:
                temp.append(cell_obj.value)
    i = 0
    while True:
        if i >= len(temp):
            break
        if temp[i] == "MD" or temp[i] == "MS" or temp[i] == "XD" or temp[i] == "WS" or temp[i] == "WD":
            curr_event = temp[i]
        elif temp[i] != str(temp[i]):
            json_rank_text = "Rank " + str(temp[i])
        else:
            name = temp[i]
            if json_rank_text not in Meet[curr_event].keys(): # if empty
                 Meet[curr_event][json_rank_text] = {"Player Name" : [name]}
            else:
                 Meet[curr_event][json_rank_text]["Player Name"].append(name)
        i+=1
def search():
    pass
def makeMeet(Meet):# for a MEET, not trimeet
    #first make structure: how long each game will take, how many courts, how much time
    #priority courts named courts 1, 2 and 3. <- these hold the best courts.
    
    print("//MAKE PROGRAM START\\\ ")
    tpm = input("what is the time given for each game cycle? write in terms of minutes(30 = 30 minutes)") #tpm = time per match
    meet_time_start = input("when does the meet start? Write in the 24:00 clock format(16:00 = 4pm)")
    meet_time_end = input("When does the meet end? Write in the 24:00 clock format(16:00 = 4pm)")
    court_number = input("How many courts are used in this meet?")
    t1 = datetime.strptime(meet_time_start, "%H:%M")#check times
    t2 = datetime.strptime(meet_time_end, "%H:%M")
    difference = t2 - t1
    plausible_minutes = difference.total_seconds() / 60 #calculates minutes alloted to the meet
    plausible_total_matches = (plausible_minutes / int(tpm)) * int(court_number) #minutes over time per match to get amount of matches for 1 court, * court no.
    MD_MAX = 0
    MS_MAX = 0
    XD_MAX = 0
    WD_MAX = 0
    WS_MAX = 0
    for i in Meet: #searches and finds the number of matches that will occur, from the highest number of spots from each event of each time.
        for j in Meet[i]:
            if j == "MD" and len(Meet[i][j].values()) > MD_MAX:
                MD_MAX = len(Meet[i][j].values())
            elif j == "MS" and len(Meet[i][j].values()) > MS_MAX:
                MS_MAX = len(Meet[i][j].values())
            elif j == "XD" and len(Meet[i][j].values()) > XD_MAX:
                XD_MAX = len(Meet[i][j].values())
            elif j == "WD" and len(Meet[i][j].values()) > WD_MAX:
                WD_MAX = len(Meet[i][j].values())
            elif j == "WS" and len(Meet[i][j].values()) > WS_MAX:
                WS_MAX = len(Meet[i][j].values())
    print("MD" , MD_MAX,"MS" , MS_MAX,"XD" , XD_MAX,"WD" , WD_MAX,"WS" , WS_MAX)
    TOTAL_MATCHES = MD_MAX + MS_MAX + XD_MAX+ WD_MAX+ WS_MAX # number of matches that have to be played
    print("PLAUSible matches", plausible_total_matches, "||TOTAL matches" , TOTAL_MATCHES)
    if(plausible_total_matches < TOTAL_MATCHES): #if amount of matches > matches can be played, impossible to create schedule
        print("NO TIME FOR THE MEET\nPlausible matches based on tpm and meet time:",plausible_total_matches, "\nTotal matches required for meet:", TOTAL_MATCHES)
        return
    



def menu():
    print("----------------------Menu----------------------\n" + " [Add Roster -> [A]] \n [View Roster -> [V]] \n [Save Changes -> [S]] \n [Make Meet -> [M]] \n [Terminate -> [X]]")
    return input("\n<Your Choice>")
def save(Meet):
    json_format = json.dumps(Meet, indent=3)

    file_to_delete = open("save.txt",'w')
    file_to_delete.close()

    with open('save.txt', 'w') as convert_file: 
        convert_file.write(json_format)
    print("saved!")
#---
teams = [] # a list of teams.
teamNumber = input("//PROGRAM START\\\ \nHow many teams are there in this meet?")
players = dict() # a dictionary of players relevant to each team.
#---
Meet = dict()
for i in range(1, int(teamNumber) + 1):
    print("--------------------------")
    temp = input("What is the team name of team " +  str(i) + "?")
    temp = temp.upper()
    Meet[temp] = dict()
    teams.append(temp)
#---- add the events within the team
addEvent(Meet)
#----
if int(teamNumber) == 2:
    print("--------------------------")
    print("Welcome to the meet between", list(Meet.keys())[0],"and", list(Meet.keys())[1] + "!")
else:
    print("--------------------------")
    print("Welcome to the Trimeet between", list(Meet.keys())[0]+",", list(Meet.keys())[1],"and", list(Meet.keys())[2] + "!")
#----gonna have to loop the process below----
while True:
    menu_query = menu()
    #try:
    if menu_query.upper() == "A":
        temp = input("Selection : Add Player : Which team is this player on?")
        if temp.upper() == list(Meet.keys())[0]:
            addPlayer(Meet[teams[0]], teams[0], players)
        elif temp.upper() == list(Meet.keys())[1]:
            addPlayer(Meet[teams[1]], teams[1], players)
        elif len(teams) > 2:
            if temp.upper() == list(Meet.keys())[2]:
                addPlayer(Meet[teams[2]], teams[2], players)
        else:
            print("Not a team!")
    if menu_query.upper() == "S":
        save(Meet)
    if menu_query.upper() == "V":
        print("this function has not been implemented yet")
        break
    if menu_query.upper() == "M":
        makeMeet(Meet)
    if menu_query.upper() == "X":
            break
    #except:
        #print("not the right answer!")# incase something is inputted wrong



#while True:
    
print(Meet)